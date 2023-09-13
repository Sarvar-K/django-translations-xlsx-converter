import os
import polib
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def get_languages(locale_path):
    return [it.name for it in os.scandir(locale_path) if it.is_dir()]


def read_po(path):
    return polib.pofile(path)


def po_to_dict(pofile):
    res = {}
    for entry in pofile:
        res[entry.msgid] = entry.msgstr
    return res


def write_messages_to_excel(excel_path, languages, language_nested_messages_dict):
    wb = Workbook()
    worksheet = wb.active
    worksheet.title = 'Переводы сообщений'

    columns = [
        ('Ключ(!!НЕ ИЗМЕНЯТЬ!!!)', 100),
    ]

    for language in languages:
        columns.append((language, 100))

    row_num = 1

    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    for message, language_dict in language_nested_messages_dict.items():
        row_num += 1

        row = [message,]

        for language in languages:
            row.append(language_dict[language])

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    wb.save(excel_path)


def get_dict_from_excel(excel_path, local_language_keys):
    res = {}
    wb = load_workbook(filename=excel_path)
    worksheet = wb.active

    rows_list = list(worksheet.iter_rows())
    excel_languages_list = [dict(
        column_index=cell.column - 1,
        name=cell.value
    ) for cell in rows_list[0][1:]]

    for row in rows_list[1:]:
        translations = _construct_translations_dict_for_key(row, excel_languages_list, local_language_keys)
        res[row[0].value] = translations

    return res


def _construct_translations_dict_for_key(row, excel_languages_list, local_language_keys):
    translations = {}

    for language in _map_excel_to_local_language(excel_languages_list, local_language_keys):
        translations[language["name"]] = row[language["column_index"]].value

    return translations


def _map_excel_to_local_language(excel_languages_list, local_language_keys):
    mapped_languages = []

    for local_language_key in local_language_keys:
        for obj in excel_languages_list:
            if local_language_key.startswith(obj["name"]):
                mapped_languages.append(dict(name=local_language_key, column_index=obj["column_index"]))

    return mapped_languages


def update_source(source_pofile, language, dict_data):
    for entry in source_pofile:
        if entry.msgid not in dict_data:
            continue

        entry.msgstr = dict_data[entry.msgid].get(language) or ""
